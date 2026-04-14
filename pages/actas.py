import streamlit as st
import pandas as pd
from supabase import create_client
from datetime import datetime
import openpyxl
import xlrd

# Configuración de página
st.set_page_config(
    page_title="Fiscalización - OSECAC",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Conexión a Supabase
SUPABASE_URL_ACTAS = st.secrets["SUPABASE_URL_ACTAS"]
SUPABASE_KEY_ACTAS = st.secrets["SUPABASE_KEY_ACTAS"]
supabase = create_client(SUPABASE_URL_ACTAS, SUPABASE_KEY_ACTAS)

# Estilo
st.markdown("""
<style>
    .main-header { background-color: #1e293b; padding: 1.2rem 1.5rem; border-radius: 8px; margin-bottom: 1.5rem; border-left: 4px solid #3b82f6; }
    .success-box { background-color: #064e3b; padding: 1rem; border-radius: 6px; border-left: 4px solid #10b981; margin: 1rem 0; }
    .warning-box { background-color: #451a03; padding: 1rem; border-radius: 6px; border-left: 4px solid #f59e0b; margin: 1rem 0; }
    .info-box { background-color: #1e293b; padding: 1rem; border-radius: 6px; border-left: 4px solid #3b82f6; margin: 1rem 0; }
    div[data-testid="stButton"] button { background-color: #3b82f6; color: white; font-weight: 500; border: none; padding: 0.4rem 1.2rem; }
    div[data-testid="stButton"] button:hover { background-color: #2563eb; }
</style>
""", unsafe_allow_html=True)

# Header
st.markdown("""
<div class="main-header">
    <h2 style="color: #ffffff; margin: 0; font-weight: 500;">Fiscalización - Deuda Presunta</h2>
    <p style="color: #94a3b8; margin: 0.3rem 0 0 0; font-size: 0.85rem;">Sistema de gestión y seguimiento</p>
</div>
""", unsafe_allow_html=True)

col_back, _ = st.columns([1, 5])
with col_back:
    if st.button("← Volver", key="btn_volver"):
        st.switch_page("main.py")

st.markdown("---")

MAPEO_EXCEL_A_TABLA = {
    'DELEGACION': 'delegacion',
    'LOCALIDAD': 'localidad',
    'CUIT': 'cuit',
    'RAZON SOCIAL': 'razon_social',
    'DEUDA PRESUNTA': 'deuda_presunta',
    'CP': 'cp',
    'CALLE': 'calle',
    'NUMERO': 'numero',
    'PISO': 'piso',
    'DPTO': 'dpto',
    'FECHARELDEPENDENCIA': 'fechareldependencia',
    'EMAIL': 'email',
    'TEL_DOM_LEGAL': 'tel_dom_legal',
    'TEL_DOM_REAL': 'tel_dom_real',
    'ULTIMA ACTA': 'ultima_acta',
    'DESDE': 'desde',
    'HASTA': 'hasta',
    'DETECTADO': 'detectado',
    'ESTADO': 'estado',
    'FECHA_PAGO_OBL': 'fecha_pago_obl',
    'EMPL 10-2025': 'empl_10_2025',
    'EMP 11-2025': 'emp_11_2025',
    'EMPL 12-2025': 'empl_12_2025',
    'ACTIVIDAD': 'actividad',
    'SITUACION': 'situacion'
}

tab1, tab2, tab3 = st.tabs(["📊 Cargar Padrón", "✏️ Editar Legajos y Vtos", "📧 Solicitar Actas"])

# ==================== TAB 1 ====================
with tab1:
    st.markdown("### Cargar Padrón de Deuda Presunta")
    
    uploaded_file = st.file_uploader(
        "Seleccionar archivo Excel",
        type=["xls", "xlsx"],
        key="upload_padron"
    )
    
    if uploaded_file is not None:
        st.info(f"Archivo: {uploaded_file.name}")
        
        try:
            # LEER EXCEL USANDO openpyxl DIRECTAMENTE (sin pandas para la lectura)
            from io import BytesIO
            
            # Cargar el workbook
            if uploaded_file.name.endswith('.xls'):
                # Para archivos .xls usamos xlrd
                workbook = xlrd.open_workbook(file_contents=uploaded_file.read())
                sheet = workbook.sheet_by_index(0)
                
                # Obtener encabezados (primera fila)
                headers = [str(sheet.cell_value(0, col)).strip().upper() for col in range(sheet.ncols)]
                
                # Obtener datos
                datos = []
                for row in range(1, sheet.nrows):
                    fila = {}
                    for col in range(sheet.ncols):
                        valor = sheet.cell_value(row, col)
                        # Convertir a string o None
                        if valor == '' or valor is None:
                            valor = None
                        else:
                            valor = str(valor).strip()
                        fila[headers[col]] = valor
                    datos.append(fila)
            else:
                # Para archivos .xlsx usamos openpyxl
                workbook = openpyxl.load_workbook(BytesIO(uploaded_file.read()), data_only=True)
                sheet = workbook.active
                
                # Obtener encabezados (primera fila)
                headers = []
                for col in range(1, sheet.max_column + 1):
                    val = sheet.cell(row=1, column=col).value
                    headers.append(str(val).strip().upper() if val else f"COL_{col}")
                
                # Obtener datos
                datos = []
                for row in range(2, sheet.max_row + 1):
                    fila = {}
                    for col in range(1, sheet.max_column + 1):
                        valor = sheet.cell(row=row, column=col).value
                        # Convertir a string o None
                        if valor is None or valor == '':
                            valor = None
                        else:
                            valor = str(valor).strip()
                        fila[headers[col-1]] = valor
                    datos.append(fila)
            
            # Ahora datos es una lista de diccionarios con TODOS los valores como string o None
            # Mapear a las columnas de la tabla
            registros_finales = []
            
            for fila in datos:
                registro = {}
                for col_excel, col_tabla in MAPEO_EXCEL_A_TABLA.items():
                    if col_excel in fila:
                        registro[col_tabla] = fila[col_excel]
                    else:
                        registro[col_tabla] = None
                
                # Agregar columnas extras
                registro['leg'] = None
                registro['vto'] = None
                registro['mail_enviado'] = 'NO'
                registro['acta'] = None
                registro['fecha_carga'] = datetime.now().isoformat()
                registro['estado_gestion'] = 'PENDIENTE'
                
                registros_finales.append(registro)
            
            st.success(f"✅ Archivo procesado: {len(registros_finales)} registros")
            
            # Mostrar vista previa
            df_preview = pd.DataFrame(registros_finales[:5])
            with st.expander("Vista previa"):
                st.dataframe(df_preview, use_container_width=True)
            
            if st.button("✅ Confirmar carga", type="primary"):
                with st.spinner("Cargando datos..."):
                    # Verificar duplicados
                    existentes = supabase.table("padron_deuda_presunta").select("cuit, ultima_acta").execute()
                    existentes_set = set()
                    for reg in existentes.data:
                        cuit = str(reg.get('cuit') or '')
                        acta = str(reg.get('ultima_acta') or '')
                        existentes_set.add((cuit, acta))
                    
                    nuevos = []
                    duplicados = 0
                    for reg in registros_finales:
                        cuit = str(reg.get('cuit') or '')
                        acta = str(reg.get('ultima_acta') or '')
                        if (cuit, acta) not in existentes_set:
                            nuevos.append(reg)
                        else:
                            duplicados += 1
                    
                    if nuevos:
                        # Insertar en lotes de 100
                        total = 0
                        for i in range(0, len(nuevos), 100):
                            lote = nuevos[i:i+100]
                            resultado = supabase.table("padron_deuda_presunta").insert(lote).execute()
                            total += len(resultado.data)
                        
                        st.success(f"✅ Carga completada: {total} registros insertados. Duplicados omitidos: {duplicados}")
                    else:
                        st.warning("Sin registros nuevos")
                            
        except Exception as e:
            st.error(f"Error: {str(e)}")

# ==================== TAB 2 ====================
with tab2:
    st.markdown("### Editar Legajos y Fechas de Vencimiento")
    
    try:
        datos = supabase.table("padron_deuda_presunta").select("*").execute()
        
        if datos.data:
            df_datos = pd.DataFrame(datos.data)
            st.write(f"**Total de registros:** {len(df_datos)}")
            
            columnas_editor = ['id', 'cuit', 'razon_social', 'leg', 'vto', 'mail_enviado', 'acta', 'estado_gestion']
            columnas_existentes = [col for col in columnas_editor if col in df_datos.columns]
            df_editable = df_datos[columnas_existentes].copy()
            
            edited_df = st.data_editor(
                df_editable,
                use_container_width=True,
                column_config={
                    "leg": st.column_config.TextColumn("LEG"),
                    "vto": st.column_config.DateColumn("VTO", format="DD/MM/YYYY"),
                    "mail_enviado": st.column_config.SelectColumn("MAIL ENVIADO", options=["NO", "SI"]),
                    "acta": st.column_config.TextColumn("ACTA"),
                    "estado_gestion": st.column_config.SelectColumn("Estado", options=["PENDIENTE", "ACTA_SOLICITADA", "ACTA_RECIBIDA", "CERRADO"]),
                },
                disabled=['id', 'cuit', 'razon_social'],
                key="editor"
            )
            
            if st.button("Guardar Cambios", type="primary"):
                with st.spinner("Guardando..."):
                    for _, row in edited_df.iterrows():
                        datos_update = {}
                        if 'leg' in edited_df.columns:
                            datos_update['leg'] = row['leg'] if pd.notna(row['leg']) and row['leg'] != '' else None
                        if 'vto' in edited_df.columns:
                            val = row['vto']
                            if pd.notna(val) and val != '':
                                if isinstance(val, (pd.Timestamp, datetime)):
                                    datos_update['vto'] = val.strftime('%Y-%m-%d')
                                else:
                                    datos_update['vto'] = str(val)
                            else:
                                datos_update['vto'] = None
                        if 'mail_enviado' in edited_df.columns:
                            datos_update['mail_enviado'] = row['mail_enviado']
                        if 'acta' in edited_df.columns:
                            datos_update['acta'] = row['acta'] if pd.notna(row['acta']) and row['acta'] != '' else None
                        if 'estado_gestion' in edited_df.columns:
                            datos_update['estado_gestion'] = row['estado_gestion']
                        
                        if datos_update:
                            supabase.table("padron_deuda_presunta").update(datos_update).eq("id", row['id']).execute()
                    
                    st.success("Cambios guardados")
                    st.rerun()
        else:
            st.info("No hay datos cargados")
    except Exception as e:
        st.error(f"Error: {str(e)}")

# ==================== TAB 3 ====================
with tab3:
    st.markdown("### Solicitar Actas a Central")
    
    try:
        datos = supabase.table("padron_deuda_presunta").select("*").execute()
        
        if datos.data:
            df_datos = pd.DataFrame(datos.data)
            
            df_listos = df_datos[
                (df_datos['leg'].notna()) & 
                (df_datos['vto'].notna()) &
                (df_datos['mail_enviado'] != 'SI')
            ]
            
            if len(df_listos) > 0:
                st.markdown(f"""
                <div class="info-box">
                    <strong>Registros listos:</strong> {len(df_listos)} empresas
                </div>
                """, unsafe_allow_html=True)
                
                st.dataframe(df_listos[['cuit', 'razon_social', 'leg', 'vto']], use_container_width=True)
                
                if st.button("Enviar solicitud", type="primary"):
                    for _, row in df_listos.iterrows():
                        supabase.table("padron_deuda_presunta").update({
                            "mail_enviado": "SI",
                            "estado_gestion": "ACTA_SOLICITADA"
                        }).eq("id", row['id']).execute()
                    
                    st.success(f"Solicitud registrada para {len(df_listos)} empresas")
            else:
                st.info("No hay registros listos")
        else:
            st.info("No hay datos cargados")
    except Exception as e:
        st.error(f"Error: {str(e)}")
